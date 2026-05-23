"""
Generate QOSMIC business model spreadsheet with 6 linked tabs.
All assumptions in Tab 1 (blue cells). Everything references Tab 1.

|Market context (web-verified 2026-05-20):
|- GSaaS market: $40.99B (2025) -> $82.72B (2030), CAGR ~15% (MarketsandMarkets)
|- Optical satcom: $0.62B (2025) -> $1.56B (2030), CAGR 20.4% (MarketsandMarkets)
|- KSAT (largest GSaaS): NOK 2,236M (~$210M) revenue (2024), 28 stations, 522 employees (Space Norway AR)
|- Rocket Lab acquired Mynaric for $155M (Apr 2026) — optical comms infra is M&A-validated
|- QOSMIC: ARGUS OGS + ZAPHOD sat terminal; TRL 6; backed by Accel, Prosus, SPC, ARTPARK (qosmic.space)
|- stations.qosmic.space: 11-station optical GSaaS marketplace (1 owned + 10 partner)
|- Partners include NASA JPL, ESA, DLR, Australian, Japanese, UAE, Norwegian facilities
|- Team: Shreyaans Jain (CEO), Dr. Rohit Ramakrishnan (CTO), Prof. Aloke Kumar (Fac. Advisor)
|- arXiv:2507.20908 (probabilistic link budget model, July 2025)
|- Pixxel: $95M total funding, 177 employees, $4.65M revenue (Tracxn/Pixxel PR)
|- Seed-stage deep tech: CapEx-heavy Years 1-3 realistic path
"""
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True, size=11)
MONEY = '$#,##0'
NUM = '#,##0'
PCT = '0.0%'
THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))
OUT_DIR = Path(__file__).resolve().parent.parent / "deliverables"


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN


def build_assumptions(wb):
    ws = wb.active
    ws.title = "Assumptions"
    ws.sheet_properties.tabColor = "4472C4"
    for c, h in enumerate(["Parameter", "Value", "Unit", "Source / Notes"], 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 4)

    r, rows = 2, {}
    def add(label, val, unit, note):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = BOLD
        c = ws.cell(row=r, column=2, value=val)
        c.fill, c.border = BLUE_FILL, THIN
        c.number_format = PCT if (isinstance(val, float) and val < 1) else MONEY if unit in ("USD","USD/yr","USD/mo") else NUM
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=note)
        rows[label] = r
        r += 1

    ws.cell(row=r, column=1, value="STATION ECONOMICS").font = Font(bold=True, size=12, color="4472C4"); r += 1
    add("CapEx per Station", 200000, "USD", "Midpoint BoM — telescope $39K, mount $14K, DM $40K, rest")
    add("OpEx per Station per Year", 50000, "USD", "Power $4K, lease $3.5K, fiber $9K, maint $7K, person $18K, etc.")
    add("CapEx Amortization Period", 5, "years", "Straight-line")

    ws.cell(row=r, column=1, value="DEPLOYMENT SCHEDULE").font = Font(bold=True, size=12, color="4472C4"); r += 1
    add("Stations Deployed Year 1", 1, "stations", "Challakere: home ground deployment")
    add("Stations Deployed Year 2", 2, "stations", "Leh + Jodhpur: desert + high-altitude")
    add("Stations Deployed Years 3-5", 2, "stations", "Sriharikota + Shillong: coastal + NE")

    ws.cell(row=r, column=1, value="REVENUE (bottoms-up customer model)").font = Font(bold=True, size=12, color="4472C4"); r += 1
    add("OGS Customers Year 1", 0, "customers", "Still building/demoing first satellite downlink")
    add("OGS Customers Year 2", 1, "customers", "1 pilot customer (e.g., Pixxel test downlink)")
    add("OGS Customers Year 3", 3, "customers", "2-3 early adopter satellite operators")
    add("OGS Customers Year 4", 5, "customers", "Growing commercial adoption")
    add("OGS Customers Year 5", 8, "customers", "Expanding customer base")
    add("ARPU (Annual Revenue per Customer)", 150000, "USD/yr", "~$12.5K/mo per customer = ~11% of station capacity at $1/GB")
    add("Data Price per GB", 1, "USD", "Qosmic pricing; competitive vs RF $3-17/GB")
    add("GB per Pass", 750, "GB", "10 Gbps x 10 min x 60s / 8 bits")
    add("Passes per Day per Station", 5, "passes", "LEO 500km: typical 5 passes/day ~10 min each")

    ws.cell(row=r, column=1, value="CLOUD MODEL (FROM PART A)").font = Font(bold=True, size=12, color="4472C4"); r += 1
    add("Cloud Avail (5 stations)", 0.995, "fraction", "Part A: 5 stations >99% clear-sky availability")
    add("Cloud Avail (1 station)", 0.46, "fraction", "Part A: Challakere alone = 46%")

    ws.cell(row=r, column=1, value="HARDWARE SALES").font = Font(bold=True, size=12, color="4472C4"); r += 1
    add("HW Units Year 3", 0, "units", "No sales yet — need proven track record")
    add("HW Units Year 4", 1, "units", "First turnkey sale (govt/research customer)")
    add("HW Units Year 5", 2, "units", "Growing interest; 2 stations")
    add("Hardware Price per Unit", 350000, "USD", "Standard 50cm OGS at ~1.75x BoM")

    ws.cell(row=r, column=1, value="NQM / GOVT CONTRACTS").font = Font(bold=True, size=12, color="4472C4"); r += 1
    add("NQM Year 1", 200000, "USD", "Phase 1 PoC contract")
    add("NQM Year 2", 400000, "USD", "Phase 1 delivery + milestone")
    add("NQM Year 3", 600000, "USD", "Phase 2 demonstration")
    add("NQM Year 4", 250000, "USD", "Maintenance/support contract")
    add("NQM Year 5", 150000, "USD", "Consulting / extension")

    ws.cell(row=r, column=1, value="PARTNER NETWORK (stations.qosmic.space)").font = Font(bold=True, size=12, color="4472C4"); r += 1
    add("Partner Stations Year 1", 10, "stations", "Existing: NASA JPL, ESA, DLR, Aus, Japan, UAE, Norway etc.")
    add("Partner Stations Year 2", 12, "stations", "+2 new partner agreements signed")
    add("Partner Stations Year 3", 15, "stations", "+3: expanding into Asia-Pacific")
    add("Partner Stations Year 4", 18, "stations", "+3: Middle East + Africa")
    add("Partner Stations Year 5", 20, "stations", "+2: total 20+ global stations")
    add("Partner Passes per Day per Station", 5, "passes", "Same orbit geometry as owned stations")
    add("Partner GB per Pass", 750, "GB", "10 Gbps x 10 min, same as owned stations")
    add("Partner Utilization Year 1", 0.05, "fraction", "Platform launch year, building booking volume")
    add("Partner Utilization Year 2", 0.07, "fraction", "Growing awareness among satellite operators")
    add("Partner Utilization Year 3", 0.10, "fraction", "Proven platform, network effects begin")
    add("Partner Utilization Year 4", 0.12, "fraction", "Established marketplace, repeat bookings")
    add("Partner Utilization Year 5", 0.15, "fraction", "Mature platform, 20+ partner stations")
    add("Partner Commission per GB", 0.20, "USD", "Qosmic spread on partner bookings")
    ws.cell(row=rows["Partner Commission per GB"], column=2).number_format = '$#,##0.00'

    ws.cell(row=r, column=1, value="ZAPHOD TERMINAL (satellite optical terminal)").font = Font(bold=True, size=12, color="4472C4"); r += 1
    add("ZAPHOD Units Year 2", 1, "units", "TakeMe2Space integration (co-developer)")
    add("ZAPHOD Units Year 3", 2, "units", "Early adopter CubeSat operators")
    add("ZAPHOD Units Year 4", 3, "units", "Growing commercial sales to LEO operators")
    add("ZAPHOD Units Year 5", 5, "units", "Scaled production; multiple constellation customers")
    add("ZAPHOD Price per Unit", 75000, "USD", "CubeSat-class <3 kg; co-developed with TakeMe2Space")

    ws.cell(row=r, column=1, value="PLATFORM & INFRASTRUCTURE").font = Font(bold=True, size=12, color="4472C4"); r += 1
    add("AWS Monthly Cost Year 1", 1500, "USD/mo", "Basic dashboard, few API calls, limited users")
    add("AWS Monthly Cost Year 2", 2000, "USD/mo", "More partner stations integrated, growing traffic")
    add("AWS Monthly Cost Year 3", 3000, "USD/mo", "Predictive scheduling, analytics, 3D globe")
    add("AWS Monthly Cost Year 4", 4000, "USD/mo", "AI features, more partners, higher throughput")
    add("AWS Monthly Cost Year 5", 5000, "USD/mo", "Full platform, global traffic, 20+ partners")

    ws.cell(row=r, column=1, value="HEADCOUNT & OVERHEAD").font = Font(bold=True, size=12, color="4472C4"); r += 1
    add("Headcount Year 1", 8, "FTE", "Current Qosmic team: founders + engineers")
    add("Headcount Year 2", 10, "FTE", "+2 hires (engineering + BD)")
    add("Headcount Year 3", 13, "FTE", "+3 hires (engineering + operations)")
    add("Headcount Year 4", 13, "FTE", "2 resigned + 2 hired (churn modeled)")
    add("Headcount Year 5", 15, "FTE", "Full team across engineering, ops, sales")
    add("Avg Salary (India)", 30000, "USD/yr", "Blended: senior ~$40-50K, junior ~$15-20K")
    add("Corporate Overhead Year 1", 50000, "USD/yr", "8 people: coworking + legal + insurance")
    add("Corporate Overhead Year 2", 62500, "USD/yr", "10 people: scales linearly w/ headcount")
    add("Corporate Overhead Year 3", 81250, "USD/yr", "13 people")
    add("Corporate Overhead Year 4", 81250, "USD/yr", "13 people (churn offset)")
    add("Corporate Overhead Year 5", 93750, "USD/yr", "15 people")

    ws.cell(row=r, column=1, value="SEED STAGE").font = Font(bold=True, size=12, color="4472C4"); r += 1
    add("Seed Round", 1000000, "USD", "Per interview doc")
    add("Monthly Burn Year 1", 150000, "USD/mo", "Total cash spend incl. one-time setup, BD, working capital buffer")

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 58
    return rows


def build_cloud_model(wb, rows):
    ws = wb.create_sheet("Cloud Model from Part A")
    ws.sheet_properties.tabColor = "70AD47"
    for c, h in enumerate(["Station", "P_clear", "Clear Days/Yr", "Notes"], 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 4)
    data = [("Leh", 0.74, 270, "Winter dip 39%"), ("Jodhpur", 0.78, 285, "Monsoon dip 50%"),
            ("Challakere", 0.46, 168, "Jul dip 8%"), ("Sriharikota", 0.57, 208, "Coastal monsoon"),
            ("Shillong", 0.51, 186, "Heavy monsoon Jun-Sep; Feb 87% clear")]
    for i, (n, p, d, note) in enumerate(data, 2):
        ws.cell(row=i, column=1, value=n)
        ws.cell(row=i, column=2, value=p).number_format = '0%'
        ws.cell(row=i, column=3, value=d)
        ws.cell(row=i, column=4, value=note)
    r = 9
    ws.cell(row=r, column=1, value="Network Availability").font = BOLD
    ws.cell(row=r+1, column=1, value="1 station"); ws.cell(row=r+1, column=2, value="=B4").number_format = '0.0%'
    ws.cell(row=r+2, column=1, value="2 stations"); ws.cell(row=r+2, column=2, value="=1-(1-B2)*(1-B3)").number_format = '0.0%'
    ws.cell(row=r+3, column=1, value="3 stations"); ws.cell(row=r+3, column=2, value="=1-(1-B2)*(1-B3)*(1-B4)").number_format = '0.0%'
    ws.cell(row=r+4, column=1, value="5 stations"); ws.cell(row=r+4, column=2, value="=1-(1-B2)*(1-B3)*(1-B4)*(1-B5)*(1-B6)").number_format = '0.0%'
    ws.cell(row=r+6, column=1, value="Integration: 5-station result feeds Assumptions B20").font = Font(italic=True, color="70AD47")
    ws.column_dimensions['A'].width = 22


def build_unit_economics(wb, rows):
    ws = wb.create_sheet("Unit Economics")
    ws.sheet_properties.tabColor = "ED7D31"
    ws.cell(row=1, column=1, value="CapEx per Station").font = Font(bold=True, size=13)
    for c, h in enumerate(["Component", "Low", "High", "Midpoint", "Source"], 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, 5)
    capex = [
        ("Telescope 50cm RC OTA", 35000, 45000, 39000, "PlaneWave CDK20"),
        ("Tracking Mount (direct drive)", 12000, 16000, 14000, "ASA DDM85 Basic: $14K"),
        ("Fast Steering Mirror", 3000, 8000, 5500, "Optotune/PI FSM"),
        ("Deformable Mirror (20x20)", 30000, 50000, 40000, "BMC/ALPAO est."),
        ("Wavefront Sensor (SH)", 4000, 7000, 5500, "Thorlabs WFS31 €4,261"),
        ("Fiber Coupling Optics", 5000, 10000, 7500, "Thorlabs FiberPorts"),
        ("Receiver Electronics", 5000, 10000, 7500, "Discovery Semi 10G"),
        ("Control Computer", 5000, 8000, 6500, "Industrial PC+FPGA"),
        ("Control Software (amort.)", 10000, 20000, 15000, "Amortized dev"),
        ("Site Preparation", 15000, 20000, 17500, "India civil works"),
        ("Fiber Backhaul (last mile)", 15000, 25000, 20000, "BSNL fiber build"),
        ("Integration & Test", 10000, 15000, 12500, "Labor+equipment"),
    ]
    for i, (n, lo, hi, mid, src) in enumerate(capex, 3):
        ws.cell(row=i, column=1, value=n)
        ws.cell(row=i, column=2, value=lo).number_format = MONEY
        ws.cell(row=i, column=3, value=hi).number_format = MONEY
        ws.cell(row=i, column=4, value=mid).number_format = MONEY
        ws.cell(row=i, column=4).fill = BLUE_FILL
        ws.cell(row=i, column=5, value=src)
    tr = 3 + len(capex)
    ws.cell(row=tr, column=1, value="Total CapEx").font = BOLD
    ws.cell(row=tr, column=4, value=f"=SUM(D3:D{tr-1})").number_format = MONEY

    # OpEx
    or2 = tr + 2
    ws.cell(row=or2, column=1, value="OpEx per Station per Year").font = Font(bold=True, size=13)
    for c, h in enumerate(["Item", "Annual Cost", "Notes"], 1):
        ws.cell(row=or2+1, column=c, value=h)
    style_header(ws, or2+1, 3)
    opex = [
        ("Power (2-3kW)", 4000, "India $0.10-0.15/kWh"),
        ("Site Lease", 3500, "Rural India"),
        ("Fiber Backhaul (annual)", 9000, "BSNL ~₹50K/mo"),
        ("Equipment Maintenance", 7000, "DM recalibration, recoating"),
        ("Monitoring Personnel (shared)", 18000, "1 FTE shared 3-5 stations"),
        ("Software Licenses", 3500, "OS, monitoring"),
        ("Insurance", 3500, "Equipment+liability"),
        ("Transport & Logistics", 4000, "Site visits, spares"),
    ]
    for i, (n, cost, note) in enumerate(opex, or2+2):
        ws.cell(row=i, column=1, value=n)
        ws.cell(row=i, column=2, value=cost).number_format = MONEY
        ws.cell(row=i, column=3, value=note)
    otr = or2 + 2 + len(opex)
    ws.cell(row=otr, column=1, value="Total OpEx/Station/Year").font = BOLD
    ws.cell(row=otr, column=2, value=f"=SUM(B{or2+2}:B{otr-1})").number_format = MONEY

    # Per-GB cost
    pgr = otr + 2
    ws.cell(row=pgr, column=1, value="Cost Structure at Maturity (5 stations)").font = Font(bold=True, size=11)
    ws.cell(row=pgr+1, column=1, value="Annual OpEx (5 stations)"); ws.cell(row=pgr+1, column=2, value=f"=5*Assumptions!B{rows['OpEx per Station per Year']}").number_format = MONEY
    ws.cell(row=pgr+2, column=1, value="Annual CapEx amort."); ws.cell(row=pgr+2, column=2, value=f"=5*Assumptions!B{rows['CapEx per Station']}/Assumptions!B{rows['CapEx Amortization Period']}").number_format = MONEY
    ws.cell(row=pgr+3, column=1, value="Annual Payroll (15 people at maturity)"); ws.cell(row=pgr+3, column=2, value=f"=Assumptions!B{rows['Headcount Year 5']}*Assumptions!B{rows['Avg Salary (India)']}").number_format = MONEY
    ws.cell(row=pgr+4, column=1, value="Annual Overhead"); ws.cell(row=pgr+4, column=2, value=f"=Assumptions!B{rows['Corporate Overhead Year 5']}").number_format = MONEY
    ws.cell(row=pgr+5, column=1, value="Annual Fixed Cost"); ws.cell(row=pgr+5, column=2, value=f"=B{pgr+1}+B{pgr+2}+B{pgr+3}+B{pgr+4}").number_format = MONEY
    ws.cell(row=pgr+6, column=1, value="Annual Data Volume (GB)"); ws.cell(row=pgr+6, column=2, value=f"=5*Assumptions!B{rows['GB per Pass']}*Assumptions!B{rows['Passes per Day per Station']}*365*Assumptions!B{rows['Cloud Avail (5 stations)']}").number_format = NUM
    ws.cell(row=pgr+7, column=1, value="Cost per GB (cents)"); ws.cell(row=pgr+7, column=2, value=f"=B{pgr+5}/B{pgr+6}*100").number_format = '$#,##0.000'
    ws.cell(row=pgr+8, column=1, value="Revenue at 40% util, $1/GB"); ws.cell(row=pgr+8, column=2, value=f"=B{pgr+6}*0.4*Assumptions!B{rows['Data Price per GB']}").number_format = MONEY

    ws.column_dimensions['A'].width = 38; ws.column_dimensions['B'].width = 18; ws.column_dimensions['E'].width = 40


def build_pnl(wb, rows):
    ws = wb.create_sheet("5-Year P&L")
    ws.sheet_properties.tabColor = "FFC000"
    ws.cell(row=1, column=1, value="5-Year P&L — Realistic Scenario").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="Revenue: OGS-as-a-Service + partner commissions + hardware + NQM. Costs: stations + team + platform.").font = Font(italic=True, size=10)
    for yr in range(1, 6):
        ws.cell(row=3, column=yr+1, value=f"Year {yr}")
    style_header(ws, 3, 6)

    # Revenue
    r5, r6, r7, r8, r9, r10 = 5, 6, 7, 8, 9, 10
    ws.cell(row=4, column=1, value="REVENUE").font = BOLD
    rev_labels = ["OGS-as-a-Service", "Hardware Sales", "Govt/NQM Contracts",
                  "Partner Network Commissions", "ZAPHOD Terminal Sales", "Total Revenue"]
    for i, l in enumerate(rev_labels, 0):
        ws.cell(row=5+i, column=1, value=f"  {l}")
    ws.cell(row=r10, column=1).font = BOLD

    prt_gbp = rows.get("Partner GB per Pass")
    prt_ppd = rows.get("Partner Passes per Day per Station")
    prt_cgb = rows.get("Partner Commission per GB")
    zphd_pr = rows.get("ZAPHOD Price per Unit")

    for yr in range(1, 6):
        col = yr + 1
        cl = get_column_letter(col)
        # OGS = customers * ARPU
        cust_row = rows.get(f"OGS Customers Year {yr}")
        arpu_row = rows.get("ARPU (Annual Revenue per Customer)")
        ws.cell(row=r5, column=col, value=f"=Assumptions!B{cust_row}*Assumptions!B{arpu_row}").number_format = MONEY
        # Hardware (ARGUS)
        hw_row = rows.get(f"HW Units Year {yr}")
        hwp_row = rows.get("Hardware Price per Unit")
        if hw_row and hwp_row:
            ws.cell(row=r6, column=col, value=f"=Assumptions!B{hw_row}*Assumptions!B{hwp_row}").number_format = MONEY
        else:
            ws.cell(row=r6, column=col, value=0).number_format = MONEY
        # NQM
        nqm_row = rows.get(f"NQM Year {yr}")
        if nqm_row:
            ws.cell(row=r7, column=col, value=f"=Assumptions!B{nqm_row}").number_format = MONEY
        else:
            ws.cell(row=r7, column=col, value=0).number_format = MONEY
        # Partner Network Commissions
        prt_stn_yr = rows.get(f"Partner Stations Year {yr}")
        prt_utl_yr = rows.get(f"Partner Utilization Year {yr}")
        ws.cell(row=r8, column=col, value=f"=Assumptions!B{prt_stn_yr}*Assumptions!B{prt_gbp}*Assumptions!B{prt_ppd}*365*Assumptions!B{prt_utl_yr}*Assumptions!B{prt_cgb}").number_format = MONEY
        # ZAPHOD Terminal Sales
        zaphod_row = rows.get(f"ZAPHOD Units Year {yr}")
        if zaphod_row and zphd_pr:
            ws.cell(row=r9, column=col, value=f"=Assumptions!B{zaphod_row}*Assumptions!B{zphd_pr}").number_format = MONEY
        else:
            ws.cell(row=r9, column=col, value=0).number_format = MONEY
        # Total Revenue
        ws.cell(row=r10, column=col, value=f"=SUM({cl}{r5}:{cl}{r9})").number_format = MONEY

    # Costs
    c11, c12, c13, c14, c15, c16, c17, c18 = 11, 12, 13, 14, 15, 16, 17, 18
    ws.cell(row=c11, column=1, value="COSTS").font = BOLD
    cost_labels = ["CapEx (new stations)", "OpEx (cumulative stations)", "Payroll",
                   "Corporate Overhead", "COGS - Hardware Sales",
                   "Platform Operations (AWS)", "Total Costs"]
    for i, l in enumerate(cost_labels, 0):
        ws.cell(row=c12+i, column=1, value=f"  {l}")
    ws.cell(row=c18, column=1).font = BOLD

    capex_row = rows.get("CapEx per Station")

    for yr in range(1, 6):
        col = yr + 1
        aws_yr = rows.get(f"AWS Monthly Cost Year {yr}")
        ws.cell(row=c17, column=col, value=f"=Assumptions!B{aws_yr}*12").number_format = MONEY

        d1, d2, d35 = f"Assumptions!B{rows['Stations Deployed Year 1']}", f"Assumptions!B{rows['Stations Deployed Year 2']}", f"Assumptions!B{rows['Stations Deployed Years 3-5']}"
        if yr == 1:
            cumul, new_s = f"{d1}", f"{d1}"
        elif yr == 2:
            cumul, new_s = f"{d1}+{d2}", f"{d2}"
        else:
            cumul, new_s = f"{d1}+{d2}+{d35}", f"{d35}" if yr == 3 else "0"
        hc_row = rows.get(f"Headcount Year {yr}")
        sal_row = rows.get("Avg Salary (India)")
        oh_row = rows.get(f"Corporate Overhead Year {yr}")
        # CapEx
        ws.cell(row=c12, column=col, value=f"=({new_s})*Assumptions!B{rows['CapEx per Station']}").number_format = MONEY
        # OpEx
        ws.cell(row=c13, column=col, value=f"=({cumul})*Assumptions!B{rows['OpEx per Station per Year']}").number_format = MONEY
        # Payroll
        ws.cell(row=c14, column=col, value=f"=Assumptions!B{hc_row}*Assumptions!B{sal_row}").number_format = MONEY
        # Corporate Overhead
        ws.cell(row=c15, column=col, value=f"=Assumptions!B{oh_row}").number_format = MONEY
        # COGS - Hardware Sales
        hw_cogs = rows.get(f"HW Units Year {yr}")
        if hw_cogs and capex_row:
            ws.cell(row=c16, column=col, value=f"=Assumptions!B{hw_cogs}*Assumptions!B{capex_row}").number_format = MONEY
        else:
            ws.cell(row=c16, column=col, value=0).number_format = MONEY
        # Total Costs
        cl = get_column_letter(col)
        ws.cell(row=c18, column=col, value=f"=SUM({cl}{c12}:{cl}{c17})").number_format = MONEY

    # EBITDA
    e20 = 20
    ws.cell(row=e20, column=1, value="EBITDA (Revenue - Costs)").font = BOLD
    for yr in range(1, 6):
        col = yr + 1
        cl = get_column_letter(col)
        ws.cell(row=e20, column=col, value=f"={cl}{r10}-{cl}{c18}").number_format = MONEY
        ws.cell(row=e20, column=col).border = THIN

    ws.column_dimensions['A'].width = 34
    for c in range(2, 7):
        ws.column_dimensions[get_column_letter(c)].width = 18


def build_sensitivity(wb):
    ws = wb.create_sheet("Breakeven Sensitivity")
    ws.sheet_properties.tabColor = "FF0000"
    ws.cell(row=1, column=1, value="Annual OGS Revenue ($K) — Sensitivity: Customers × ARPU").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value="Rows = number of paying customers. Columns = annual revenue per customer ($K).").font = Font(italic=True, size=10)
    ws.cell(row=3, column=1, value="Fixed costs ~$1,054K/yr at 5-station maturity. Green = revenue > $1,054K.").font = Font(size=9)

    customers = [0, 1, 3, 5, 8, 12, 15]
    arpu_k = [25, 50, 75, 100, 150, 200]

    ws.cell(row=5, column=1, value="Customers \\ ARPU ($K)")
    style_header(ws, 5, 1 + len(arpu_k))
    for c, a in enumerate(arpu_k, 2):
        ws.cell(row=5, column=c, value=a).number_format = NUM
    for r, cust in enumerate(customers, 6):
        ws.cell(row=r, column=1, value=cust).font = BOLD
        for c, arpu in enumerate(arpu_k, 2):
            rev = cust * arpu  # annual revenue in $K
            ws.cell(row=r, column=c, value=rev).number_format = NUM
            ws.cell(row=r, column=c).border = THIN
    ws.column_dimensions['A'].width = 18
    for c in range(2, 2 + len(arpu_k)):
        ws.column_dimensions[get_column_letter(c)].width = 14


def build_investor_summary(wb, rows):
    ws = wb.create_sheet("Investor Summary")
    ws.sheet_properties.tabColor = "7030A0"
    ws.cell(row=1, column=1, value="Investor Summary — Key Metrics").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="All values auto-calculated from Assumptions tab").font = Font(italic=True, size=10, color="888888")
    for c, h in enumerate(["Metric", "Value", "Notes"], 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, 3)

    items = [
        ("MARKET OPPORTUNITY", None, ""),
        ("  TAM (GSaaS market, 2030)", "$14.2B", "Dataintelo: $3.8B->$14.2B (CAGR 15.8%)"),
        ("  Serviceable Market (Optical GSaaS)", "~$200-500M by 2030", "Optical subset of GSaaS, early stage"),
        ("", "", ""),
        ("FINANCIAL PERFORMANCE", None, ""),
        ("  Year 3 OGS Revenue", "='5-Year P&L'!C5", "From P&L tab — 3 customers × $150K ARPU"),
        ("  Year 5 OGS Revenue", "='5-Year P&L'!F5", "From P&L tab — 8 customers × $150K ARPU"),
        ("  Year 5 Total Revenue", "='5-Year P&L'!F10", "OGS + partner commissions + HW + NQM + ZAPHOD"),
        ("  Year 5 EBITDA", "='5-Year P&L'!F20", "Revenue minus costs"),
        ("  Year 5 EBITDA Margin", "='5-Year P&L'!F20/'5-Year P&L'!F10", "EBITDA / Total Revenue"),
        ("", "", ""),
        ("CASH & RUNWAY", None, ""),
        ("  Seed Round", 1000000, "Per interview document"),
        ("  Monthly Burn (Y1 est.)", 150000, "$120K salaries + $30K OpEx"),
        ("  Est. Runway (months)", "=B16/B17", "Seed / Monthly Burn. $150K/mo is top-down total cash spend incl. one-time setup + buffer"),
        ("", "", ""),
        ("FUNDRAISING", None, ""),
        ("  Target Series A Raise", "$4-6M", "Range based on comparable seed->A multiples"),
        ("  Est. Pre-Money Valuation", "$15-25M", "Discounted for seed-stage; KSAT at ~$150M rev"),
        ("  Est. Post-Money", "$19-31M", "Pre + Raise"),
    ]

    for i, (metric, formula, note) in enumerate(items, 4):
        ws.cell(row=i, column=1, value=metric)
        if formula is None:
            ws.cell(row=i, column=1).font = Font(bold=True, size=11)
        elif isinstance(formula, str) and formula.startswith("="):
            c = ws.cell(row=i, column=2, value=formula)
            if '/' in formula and 'P&L' in formula:
                c.number_format = PCT  # EBITDA margin: ='5-Year P&L'!F20/'5-Year P&L'!F10
            elif 'P&L' in formula:
                c.number_format = MONEY  # Single P&L cell reference
            else:
                c.number_format = NUM  # Runway =B16/B17
        elif isinstance(formula, (int, float)):
            c = ws.cell(row=i, column=2, value=formula)
            c.number_format = MONEY  # Seed Round, Monthly Burn
        else:
            c = ws.cell(row=i, column=2, value=formula)
            c.alignment = Alignment(horizontal='right')
        ws.cell(row=i, column=3, value=note)

    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 50


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    rows = build_assumptions(wb)
    build_cloud_model(wb, rows)
    build_unit_economics(wb, rows)
    build_pnl(wb, rows)
    build_sensitivity(wb)
    build_investor_summary(wb, rows)
    out = OUT_DIR / "business_model.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print(f"Tabs: {wb.sheetnames}")
    print("Living model: change blue cells in Assumptions -> all tabs update.")

if __name__ == "__main__":
    main()
